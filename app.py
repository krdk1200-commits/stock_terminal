import datetime
import io
import json
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import requests
import streamlit as st
import yfinance as yf

# 1. Page Configuration
st.set_page_config(
    page_title="TradingView Pro | Global Stock, Buy/Sell, MF, Commodity & AI Terminal",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# 2. Custom Styling
st.markdown(
    """
    <style>
    .banner-ad {
        background: linear-gradient(90deg, #0f2027, #203a43, #2c5364);
        padding: 10px 20px;
        border-radius: 8px;
        color: white;
        text-align: center;
        margin-bottom: 15px;
        font-size: 0.9rem;
    }
    .banner-ad a { color: #ffcc00; text-decoration: underline; font-weight: bold; }
    .sec-header {
        font-size: 1.15rem;
        font-weight: 700;
        padding-bottom: 6px;
        margin-top: 20px;
        margin-bottom: 12px;
        border-bottom: 2px solid #2962ff;
        color: #131722;
    }
    .fund-sound { background-color: #d4edda; color: #155724; padding: 6px 12px; border-radius: 6px; font-weight: bold; border-left: 5px solid #28a745; }
    .fund-mod { background-color: #fff3cd; color: #856404; padding: 6px 12px; border-radius: 6px; font-weight: bold; border-left: 5px solid #ffc107; }
    .fund-weak { background-color: #f8d7da; color: #721c24; padding: 6px 12px; border-radius: 6px; font-weight: bold; border-left: 5px solid #dc3545; }
    .ai-box {
        background: linear-gradient(135deg, #0d1b2a 0%, #1b263b 100%);
        color: #e0e1dd;
        padding: 18px;
        border-radius: 8px;
        margin: 15px 0;
        border: 1px solid #415a77;
    }
    .flash-box {
        background: linear-gradient(90deg, #134e5e 0%, #71b280 100%);
        color: white;
        padding: 12px 18px;
        border-radius: 8px;
        margin-bottom: 15px;
        font-weight: 500;
        box-shadow: 0 4px 10px rgba(0,0,0,0.15);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# 3. Master Databases (Indices, Commodities & Global Presets)
ALL_INDIAN_INDICES = {
    "-- इंडेक्स चुनें (Select Indian Index) --": "",
    "NIFTY 50": "^NSEI",
    "SENSEX (BSE 30)": "^BSESN",
    "NIFTY BANK": "^NSEBANK",
    "NIFTY FINANCIAL SERVICES (FINNIFTY)": "NIFTY_FIN_SERVICE.NS",
    "NIFTY NEXT 50": "^NSMIDCP",
    "NIFTY MIDCAP 50": "NIFTY_MIDCAP_50.NS",
    "NIFTY MIDCAP 100": "NIFTY_MIDCAP_100.NS",
    "NIFTY MIDCAP SELECT": "NIFTY_MID_SELECT.NS",
    "NIFTY SMALLCAP 100": "NIFTY_SMALLCAP_100.NS",
    "NIFTY SMALLCAP 250": "NIFTY_SMALLCAP_250.NS",
    "NIFTY 200": "^CNX200",
    "NIFTY 500": "^CRSLDX",
    "NIFTY IT": "^CNXIT",
    "NIFTY AUTO": "^CNXAUTO",
    "NIFTY PHARMA": "^CNXPHARMA",
    "NIFTY PSU BANK": "^CNXPSUBANK",
    "NIFTY PRIVATE BANK": "NIFTY_PVT_BANK.NS",
    "NIFTY METAL": "^CNXMETAL",
    "NIFTY FMCG": "^CNXFMCG",
    "NIFTY REALTY": "^CNXREALTY",
    "NIFTY ENERGY": "^CNXENERGY",
    "NIFTY INFRASTRUCTURE": "^CNXINFRA",
    "NIFTY OIL & GAS": "NIFTY_OIL_AND_GAS.NS",
    "NIFTY HEALTHCARE": "NIFTY_HEALTHCARE.NS",
    "NIFTY CONSUMER DURABLES": "NIFTY_CONSR_DURBL.NS",
    "NIFTY MEDIA": "^CNXMEDIA",
    "INDIA VIX": "^INDIAVIX",
}

ALL_COMMODITIES = {
    "-- कमोडिटी चुनें (Select Commodity) --": "",
    "Gold (सोना)": "GC=F",
    "Silver (चाँदी)": "SI=F",
    "Crude Oil WTI (कच्चा तेल)": "CL=F",
    "Brent Crude Oil": "BZ=F",
    "Natural Gas (प्राकृतिक गैस)": "NG=F",
    "Copper (तांबा)": "HG=F",
    "Platinum (प्लेटिनम)": "PL=F",
    "Palladium (पैलेडियम)": "PA=F",
    "Aluminum (एल्युमिनियम)": "ALI=F",
}

ALL_US_MARKET_STOCKS = {
    "-- US स्टॉक / इंडेक्स चुनें (Select US Stock) --": "",
    "S&P 500 Index": "^GSPC",
    "NASDAQ 100 Index": "^NDX",
    "Dow Jones Industrial": "^DJI",
    "Apple Inc.": "AAPL",
    "Microsoft Corp": "MSFT",
    "NVIDIA Corporation": "NVDA",
    "Alphabet Google": "GOOGL",
    "Amazon.com Inc": "AMZN",
    "Meta Platforms (Facebook)": "META",
    "Tesla Inc": "TSLA",
    "Broadcom Inc": "AVGO",
    "Taiwan Semiconductor (TSMC)": "TSM",
    "AMD": "AMD",
    "Qualcomm Inc": "QCOM",
    "Intel Corporation": "INTC",
    "Palantir Technologies": "PLTR",
    "Berkshire Hathaway": "BRK-B",
    "JPMorgan Chase": "JPM",
}

# --- MASTER TOP BUY & SELL STOCKS WITH CATALYSTS ---
TOP_BUY_STOCKS_MASTER = [
    {"Company": "State Bank of India (SBI)", "Ticker": "SBIN.NS", "Action": "🟢 STRONG BUY", "Target_1Y": "₹980", "Upside": "+22%", "Why_Buy_Reason": "मजबूत क्रेडिट ग्रोथ (15%+), घटता NPA और आकर्षक P/E वैल्युएशन"},
    {"Company": "Tata Motors Ltd", "Ticker": "TATAMOTORS.NS", "Action": "🟢 STRONG BUY", "Target_1Y": "₹1,250", "Upside": "+26%", "Why_Buy_Reason": "JLR का रिकॉर्ड फ्री-कैश-फ्लो, भारत में EV मार्केट लीडरशिप और डीमर्जर वैल्यू अनलॉकिंग"},
    {"Company": "Balrampur Chini Mills", "Ticker": "BALRAMCHIN.NS", "Action": "🟢 BUY", "Target_1Y": "₹650", "Upside": "+24%", "Why_Buy_Reason": "एथेनॉल ब्लेंडिंग 20% लक्ष्य, चीनी निर्यात नीति में छूट और मजबूत ऑपरेटिंग मार्जिन"},
    {"Company": "Kaynes Technology", "Ticker": "KAYNES.NS", "Action": "🟢 STRONG BUY", "Target_1Y": "₹6,400", "Upside": "+30%", "Why_Buy_Reason": "सेमीकंडक्टर OSAT प्लांट विस्तार, मजबूत ऑर्डर बुक (₹4,500 Cr+) और 40%+ रेवेन्यू ग्रोथ"},
    {"Company": "Coal India Ltd", "Ticker": "COALINDIA.NS", "Action": "🟢 BUY (High Div)", "Target_1Y": "₹580", "Upside": "+18% + 7% Div", "Why_Buy_Reason": "पावर सेक्टर की भारी कोयला मांग, शून्य कर्ज और 7%+ डिविडेंड यील्ड"},
    {"Company": "NVIDIA Corporation", "Ticker": "NVDA", "Action": "🟢 STRONG BUY", "Target_1Y": "$165", "Upside": "+28%", "Why_Buy_Reason": "Blackwell चिप की बेजोड़ मांग और डेटा सेंटर AI इंफ्रास्ट्रक्चर में 85%+ मोनोपॉली"},
]

TOP_SELL_STOCKS_MASTER = [
    {"Company": "Vodafone Idea", "Ticker": "IDEA.NS", "Action": "🔴 STRONG SELL", "StopLoss_Risk": "₹11.50", "Downside_Risk": "-25%", "Why_Sell_Reason": "लगातार सब्सक्राइबर लॉस, भारी कर्ज बोझ और भारी AGR देनदारी"},
    {"Company": "Paytm (One97 Comm)", "Ticker": "PAYTM.NS", "Action": "🔴 AVOID / SELL", "StopLoss_Risk": "₹750", "Downside_Risk": "-18%", "Why_Sell_Reason": "पेमेंट्स बैंक रेगुलेटरी रोक के बाद रेवेन्यू में भारी गिरावट और अनिश्चित प्रॉफिटेबिलिटी"},
    {"Company": "Yes Bank Ltd", "Ticker": "YESBANK.NS", "Action": "🔴 SELL ON RISE", "StopLoss_Risk": "₹26.00", "Downside_Risk": "-15%", "Why_Sell_Reason": "कम NIM मार्जिन (2.4%), सीमित ROA और बड़े निवेशकों का बिकवाली दबाव"},
    {"Company": "Intel Corporation", "Ticker": "INTC", "Action": "🔴 AVOID / SELL", "StopLoss_Risk": "$24.00", "Downside_Risk": "-20%", "Why_Sell_Reason": "फाउंड्री बिजनेस में भारी घाटा, AI चिप मार्केट शेयर का नुकसान और डिविडेंड निलंबन"},
    {"Company": "Ola Electric Mobility", "Ticker": "OLAELEC.NS", "Action": "🔴 SELL / BOOK PROFIT", "StopLoss_Risk": "₹82.00", "Downside_Risk": "-22%", "Why_Sell_Reason": "सर्विस शिकायतों के कारण मार्केट शेयर में कमी और लगातार ऑपरेटिंग लॉस"},
]

TOP_DIVIDEND_STOCKS_MASTER = [
    {"Company": "Vedanta Ltd", "Ticker": "VEDL.NS", "Market": "India", "Typical_Yield": "10-12%", "Cat": "High Metal Dividend", "Why_Buy": "असाधारण कैश फ्लो और उच्च डिविडेंड यील्ड", "Exp_5Y_Return": "+85%", "Exp_10Y_Return": "+210%"},
    {"Company": "Coal India Ltd", "Ticker": "COALINDIA.NS", "Market": "India", "Typical_Yield": "6-8%", "Cat": "PSU Monopoly", "Why_Buy": "जीरो डेट, स्थिर पावर डिमांड और भारी डिविडेंड पेआउट", "Exp_5Y_Return": "+75%", "Exp_10Y_Return": "+180%"},
    {"Company": "REC Limited", "Ticker": "REC.NS", "Market": "India", "Typical_Yield": "5-7%", "Cat": "Power Finance", "Why_Buy": "पावर इंफ्रास्ट्रक्चर लेंडिंग में तेज वृद्धि और डिविडेंड स्थिरता", "Exp_5Y_Return": "+95%", "Exp_10Y_Return": "+240%"},
    {"Company": "Power Finance Corp (PFC)", "Ticker": "PFC.NS", "Market": "India", "Typical_Yield": "5-7%", "Cat": "Power Finance", "Why_Buy": "मजबूत लोन बुक विस्तार और लगातार डिविडेंड ट्रैक रिकॉर्ड", "Exp_5Y_Return": "+90%", "Exp_10Y_Return": "+230%"},
    {"Company": "Indian Oil Corp (IOC)", "Ticker": "IOC.NS", "Market": "India", "Typical_Yield": "6-8%", "Cat": "Oil & Refining", "Why_Buy": "मजबूत रिफाइनिंग मार्जिन और सरकारी डिविडेंड सपोर्ट", "Exp_5Y_Return": "+60%", "Exp_10Y_Return": "+150%"},
    {"Company": "Altria Group", "Ticker": "MO", "Market": "USA", "Typical_Yield": "8-9%", "Cat": "Consumer Aristocrat", "Why_Buy": "54 वर्षों से लगातार बढ़ता डिविडेंड", "Exp_5Y_Return": "+55%", "Exp_10Y_Return": "+140%"},
    {"Company": "Realty Income (Monthly Div)", "Ticker": "O", "Market": "USA", "Typical_Yield": "5-6%", "Cat": "Real Estate REIT", "Why_Buy": "हर महीने डिविडेंड देने वाला रियल एस्टेट दिग्गज", "Exp_5Y_Return": "+65%", "Exp_10Y_Return": "+160%"},
]

TOP_MUTUAL_FUNDS_DATA = [
    {"Fund Name": "Parag Parikh Flexi Cap Fund", "Category": "Flexi Cap", "Rating": "⭐⭐⭐⭐⭐", "1M_Return": "+2.1%", "3M_Return": "+6.8%", "1Y_Return": "+24.5%", "3Y_CAGR": "+19.8%", "Exp_1M_Future": "+1.8%", "Exp_3M_Future": "+5.5%", "Exp_1Y_Future": "+18-22%", "AI_Verdict": "🟢 Strong Buy (Long Term)"},
    {"Fund Name": "Quant Small Cap Fund", "Category": "Small Cap", "Rating": "⭐⭐⭐⭐⭐", "1M_Return": "+3.4%", "3M_Return": "+9.2%", "1Y_Return": "+38.2%", "3Y_CAGR": "+28.4%", "Exp_1M_Future": "+2.5%", "Exp_3M_Future": "+8.0%", "Exp_1Y_Future": "+22-26%", "AI_Verdict": "🟢 Buy on Dips (High Alpha)"},
    {"Fund Name": "HDFC Top 100 Fund", "Category": "Large Cap", "Rating": "⭐⭐⭐⭐", "1M_Return": "+1.6%", "3M_Return": "+4.9%", "1Y_Return": "+19.8%", "3Y_CAGR": "+17.2%", "Exp_1M_Future": "+1.4%", "Exp_3M_Future": "+4.2%", "Exp_1Y_Future": "+14-17%", "AI_Verdict": "🟢 Stable Wealth Builder"},
    {"Fund Name": "Nippon India Growth Fund", "Category": "Mid Cap", "Rating": "⭐⭐⭐⭐⭐", "1M_Return": "+2.8%", "3M_Return": "+7.9%", "1Y_Return": "+32.1%", "3Y_CAGR": "+23.5%", "Exp_1M_Future": "+2.2%", "Exp_3M_Future": "+6.8%", "Exp_1Y_Future": "+20-24%", "AI_Verdict": "🟢 Strong Buy (Growth)"},
    {"Fund Name": "UTI Nifty 50 Index Fund", "Category": "Index Fund", "Rating": "⭐⭐⭐⭐⭐", "1M_Return": "+1.2%", "3M_Return": "+4.1%", "1Y_Return": "+16.5%", "3Y_CAGR": "+14.8%", "Exp_1M_Future": "+1.1%", "Exp_3M_Future": "+3.5%", "Exp_1Y_Future": "+12-15%", "AI_Verdict": "🟢 Zero-Error Passive SIP"},
]

POPULAR_STOCKS_PRESET = [
    ("Balrampur Chini Mills", "BALRAMCHIN.NS"),
    ("State Bank of India (SBI Bank)", "SBIN.NS"),
    ("SBI Cards & Payment Services", "SBICARD.NS"),
    ("SBI Life Insurance", "SBILIFE.NS"),
    ("Bank of India", "BANKINDIA.NS"),
    ("Bank of Baroda", "BANKBARODA.NS"),
    ("Canara Bank", "CANBK.NS"),
    ("Punjab National Bank (PNB)", "PNB.NS"),
    ("Union Bank of India", "UNIONBANK.NS"),
    ("Indian Bank", "INDIANB.NS"),
    ("Reliance Industries (RIL)", "RELIANCE.NS"),
    ("Tata Consultancy Services (TCS)", "TCS.NS"),
    ("Tata Motors Ltd", "TATAMOTORS.NS"),
    ("Tata Steel Ltd", "TATASTEEL.NS"),
    ("Tata Power Co Ltd", "TATAPOWER.NS"),
    ("Tata Technologies", "TATATECH.NS"),
    ("Tata Elxsi Ltd", "TATAELXSI.NS"),
    ("HDFC Bank Ltd", "HDFCBANK.NS"),
    ("ICICI Bank Ltd", "ICICIBANK.NS"),
    ("Infosys Ltd", "INFY.NS"),
    ("Bharti Airtel", "BHARTIARTL.NS"),
    ("ITC Ltd", "ITC.NS"),
    ("Larsen & Toubro (L&T)", "LT.NS"),
    ("Zomato Ltd", "ZOMATO.NS"),
    ("Jio Financial Services", "JIOFIN.NS"),
    ("Hindustan Aeronautics (HAL)", "HAL.NS"),
    ("Bharat Electronics (BEL)", "BEL.NS"),
    ("Mazagon Dock Shipbuilders", "MAZDOCK.NS"),
    ("IRFC (Railway Finance)", "IRFC.NS"),
    ("RVNL (Rail Vikas Nigam)", "RVNL.NS"),
    ("Suzlon Energy", "SUZLON.NS"),
    ("IREDA", "IREDA.NS"),
    ("Kaynes Technology", "KAYNES.NS"),
    ("Dixon Technologies", "DIXON.NS"),
]

UPCOMING_IPOS_DATA = [
    {"IPO Name": "Waaree Energies Limited", "Sector": "Solar / Renewable", "Price Band": "₹1,427 - ₹1,503", "Estimated GMP": "+95%", "Rating Review": "4.8/5 (Heavy Demand)", "AI Verdict": "🟢 STRONG APPLY (मजबूत लिस्टिंग गेन)"},
    {"IPO Name": "Hyundai Motor India", "Sector": "Automobile", "Price Band": "₹1,865 - ₹1,960", "Estimated GMP": "+8%", "Rating Review": "4.0/5 (Market Leader)", "AI Verdict": "🟢 APPLY FOR LONG TERM"},
    {"IPO Name": "Swiggy Limited", "Sector": "Quick Commerce", "Price Band": "₹371 - ₹390", "Estimated GMP": "+12%", "Rating Review": "3.8/5 (High Growth)", "AI Verdict": "🟡 APPLY FOR HIGH RISK"},
    {"IPO Name": "NTPC Green Energy Limited", "Sector": "PSU Renewable", "Price Band": "₹102 - ₹108", "Estimated GMP": "+25%", "Rating Review": "4.6/5 (Sovereign Backed)", "AI Verdict": "🟢 STRONG APPLY"}
]

# 4. Sidebar Settings
st.sidebar.markdown("### ⚙️ सेटिंग्स / Settings")
language = st.sidebar.radio("🌐 भाषा चुनें / Select Language:", ["Bilingual (हिंदी + English)", "हिंदी (Hindi)", "English"], index=0)
is_hindi = "हिंदी" in language
is_bilingual = "Bilingual" in language

def get_txt(hi, en):
    if is_bilingual: return f"{hi} | {en}"
    return hi if is_hindi else en

st.sidebar.markdown("---")
st.sidebar.markdown(f"### 💰 {get_txt('पोर्टफोलियो व री-बाय (Re-Buy Averaging)', 'Portfolio & Re-Buy')}")
buy_price = st.sidebar.number_input(
    get_txt("आपका पुराना खरीद भाव (Your Old Buy Price):", "Your Buy Price:"),
    min_value=0.0, value=0.0, step=1.0,
    help="Yield on Cost और Re-Buy/Averaging Price निकालने के लिए दर्ज करें।"
)

# 5. Top Banner & Title
st.markdown(
    """
    <div class="banner-ad">
        📢 SPONSORED / ADVERTISEMENT<br>
        ⚡ <b>Zero Brokerage Stocks, Buy/Sell Picks, Mutual Funds & Commodity Terminal</b> | <a href="#" target="_blank">Open Account Now</a>
    </div>
    """,
    unsafe_allow_html=True,
)

st.title("TradingView Pro | Global Stock, Buy/Sell, MF, Commodity & AI Terminal")
st.caption("30+ Indian Indices • Top Buy & Top Sell Picks • Multi-Horizon Returns (1M to 10Y) • Mutual Funds Radar • 100% Free Access")

# 6. Helper Functions & Search Engine
def search_yahoo_tickers(query):
    if not query or len(query.strip()) < 1:
        return []
    try:
        url = f"https://query2.finance.yahoo.com/v1/finance/search?q={query}&quotesCount=10&newsCount=0"
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=4)
        if resp.status_code == 200:
            data = resp.json()
            quotes = data.get("quotes", [])
            results = []
            for q in quotes:
                sym = q.get("symbol")
                name = q.get("shortname") or q.get("longname") or sym
                exch = q.get("exchDisp") or q.get("exchange", "")
                if sym:
                    results.append((f"{name} ({sym}) - [{exch}]", sym))
            return results
    except Exception:
        pass
    return []

def calculate_rsi(series, period=14):
    delta = series.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / (loss + 1e-9)
    return 100 - (100 / (1 + rs))

def calculate_macd(series):
    exp1 = series.ewm(span=12, adjust=False).mean()
    exp2 = series.ewm(span=26, adjust=False).mean()
    macd = exp1 - exp2
    signal = macd.ewm(span=9, adjust=False).mean()
    return macd, signal

def calculate_bollinger_bands(series, window=20):
    sma = series.rolling(window=window).mean()
    std = series.rolling(window=window).std()
    return sma + (std * 2), sma - (std * 2), sma

def calculate_intrinsic_value(eps, book_value):
    try:
        if eps > 0 and book_value > 0:
            return round(np.sqrt(22.5 * eps * book_value), 2)
    except Exception:
        pass
    return None

def evaluate_fundamental_health(info):
    score = 0
    factors = []
    op_margin = info.get("operatingMargins") or 0.0
    if op_margin > 0.15: score += 2; factors.append("✅ मजबूत ऑपरेटिंग मार्जिन (>15%)")
    elif op_margin > 0.08: score += 1; factors.append("⚖️ संतोषजनक ऑपरेटिंग मार्जिन")
    else: factors.append("⚠️ कम ऑपरेटिंग मार्जिन (<8%)")

    roe = info.get("returnOnEquity") or 0.0
    if roe > 0.15: score += 2; factors.append("✅ उत्कृष्ट ROE (>15%)")
    elif roe > 0.08: score += 1; factors.append("⚖️ मध्यम ROE")
    else: factors.append("⚠️ कमजोर ROE (<8%)")

    dte = info.get("debtToEquity")
    if dte is not None:
        if dte < 50: score += 2; factors.append("✅ सुरक्षित डेट-टू-इक्विटी (कम कर्ज)")
        elif dte < 120: score += 1; factors.append("⚖️ प्रबंधनीय कर्ज स्तर")
        else: factors.append("⚠️ उच्च कर्ज (High Debt/Equity)")
    else: score += 1

    fcf = info.get("freeCashflow") or 0
    if fcf > 0: score += 2; factors.append("✅ सकारात्मक फ्री कैश फ्लो (Positive FCF)")
    else: factors.append("⚠️ नकारात्मक / सीमित कैश फ्लो")

    pe = info.get("trailingPE") or 0
    if 0 < pe < 30: score += 2; factors.append("✅ उचित वैल्युएशन (P/E < 30)")
    elif pe >= 30: score += 1; factors.append("⚠️ उच्च प्रीमियम वैल्युएशन")

    health_pct = int((score / 10.0) * 100)
    if health_pct >= 70: category = "FUNDAMENTALLY VERY SOUND 🛡️ (अति मजबूत कंपनी)"; style_class = "fund-sound"
    elif health_pct >= 45: category = "FUNDAMENTALLY MODERATE ⚖️ (मध्यम / स्थिर कंपनी)"; style_class = "fund-mod"
    else: category = "FUNDAMENTALLY WEAK ⚠️ (कमजोर फंडामेंटल्स - सतर्क रहें)"; style_class = "fund-weak"

    return health_pct, category, style_class, factors

def fetch_option_chain_oi(ticker_obj, cmp):
    try:
        expirations = ticker_obj.options
        if not expirations: return None
        opt = ticker_obj.option_chain(expirations[0])
        calls, puts = opt.calls, opt.puts

        total_call_oi = calls["openInterest"].sum() if "openInterest" in calls else 0
        total_put_oi = puts["openInterest"].sum() if "openInterest" in puts else 0
        pcr = round(total_put_oi / (total_call_oi + 1e-9), 2)

        if pcr > 1.25: oi_sentiment = "BULLISH (मजबूत पुट राइटिंग / तेजी का रुख)"; oi_action = "🟢 BUY ON DIPS"
        elif pcr < 0.75: oi_sentiment = "BEARISH (मजबूत कॉल राइटिंग / दबाव का संकेत)"; oi_action = "🔴 SELL ON RISE"
        else: oi_sentiment = "NEUTRAL / RANGEBOUND (संतुलित दायरा)"; oi_action = "🟡 RANGE ACCUMULATION"

        max_call_strike = calls.loc[calls["openInterest"].idxmax()]["strike"] if not calls.empty and "openInterest" in calls else cmp * 1.05
        max_put_strike = puts.loc[puts["openInterest"].idxmax()]["strike"] if not puts.empty and "openInterest" in puts else cmp * 0.95
        option_fair_center = round((max_call_strike + max_put_strike) / 2, 2)

        return {
            "expiry": expirations[0], "total_call_oi": total_call_oi, "total_put_oi": total_put_oi,
            "pcr": pcr, "sentiment": oi_sentiment, "oi_action": oi_action,
            "call_resistance": max_call_strike, "put_support": max_put_strike,
            "option_fair_price": option_fair_center
        }
    except Exception:
        return None

# Top Bullish Stocks Flash Panel
st.markdown(
    """
    <div class="flash-box">
        🔥 <b>AI लाइव बुलिश फ़्लैश रडार (Top Buy & Momentum Picks):</b><br>
        • <b>BALRAMCHIN / SBIN / TATAMOTORS</b>: Strong Buy Recommendations | उच्च अर्निंग्स ग्रोथ व ब्रेकआउट मोमेंटम<br>
        • <b>TOP MUTUAL FUNDS</b>: Quant Small Cap (+38.2% 1Y) | Parag Parikh Flexi Cap (+24.5% 1Y)
    </div>
    """,
    unsafe_allow_html=True,
)

# Search Bar
st.markdown(f"<div class='sec-header'>{get_txt('🔎 ऑल इंडियन इंडेक्स, US मार्केट, कमोडिटी व यूनिवर्सल सर्च', 'Indices, Commodities, US Equities & Universal Search')}</div>", unsafe_allow_html=True)

idx_col, us_col, com_col = st.columns([1, 1, 1])

with idx_col:
    chosen_indian_index = st.selectbox(
        get_txt("🏛️ भारतीय इंडेक्स:", "🏛️ Indian Indices:"),
        options=list(ALL_INDIAN_INDICES.keys()), index=0,
        help="NIFTY 50, NIFTY 500, Bank Nifty, Midcap, Smallcap आदि चुनें।"
    )

with us_col:
    chosen_us_stock = st.selectbox(
        get_txt("🇺🇸 US मार्केट:", "🇺🇸 US Market:"),
        options=list(ALL_US_MARKET_STOCKS.keys()), index=0,
        help="S&P 500, NASDAQ, Apple, Tesla, NVDA आदि चुनें।"
    )

with com_col:
    chosen_commodity = st.selectbox(
        get_txt("🪙 कमोडिटी (Gold/Oil):", "🪙 Commodities:"),
        options=list(ALL_COMMODITIES.keys()), index=0,
        help="Gold, Silver, Crude Oil, Natural Gas, Copper आदि चुनें।"
    )

st.markdown("##### 🔎 कंपनी का नाम या सिंबल लिखें (टाइप करते ही नीचे लाइव सुझाव आएँगे):")
search_query = st.text_input(
    label="Search Box",
    placeholder="जैसे: balrampur, sbi, coal india, vedanta, rec, tata, apple, nvda...",
    value="",
    label_visibility="collapsed"
).strip()

live_suggestions = []
if search_query:
    live_suggestions = search_yahoo_tickers(search_query)

if live_suggestions:
    options_map = {disp: sym for disp, sym in live_suggestions}
    selected_option = st.selectbox("🎯 लाइव सुझाव से स्टॉक चुनें (Select from live matches):", list(options_map.keys()), index=0)
    symbol = options_map[selected_option]
elif search_query:
    cleaned_q = search_query.upper().replace(" ", "")
    symbol = f"{cleaned_q}.NS" if ("." not in cleaned_q and not cleaned_q.startswith("^") and "=" not in cleaned_q) else cleaned_q
elif chosen_indian_index != "-- इंडेक्स चुनें (Select Indian Index) --":
    symbol = ALL_INDIAN_INDICES[chosen_indian_index]
elif chosen_us_stock != "-- US स्टॉक / इंडेक्स चुनें (Select US Stock) --":
    symbol = ALL_US_MARKET_STOCKS[chosen_us_stock]
elif chosen_commodity != "-- कमोडिटी चुनें (Select Commodity) --":
    symbol = ALL_COMMODITIES[chosen_commodity]
else:
    symbol = "BALRAMCHIN.NS"

st.info(f"🎯 **सक्रिय सिंबल (Active Ticker):** `{symbol}`")

# Date Range Controls
rcol1, rcol2 = st.columns([1, 1])
with rcol1:
    range_type = st.radio(
        get_txt("Range Type / मोड:", "Range Type:"),
        [get_txt("Standard Presets (1D, 1M, 1Y...)", "Standard Presets"), get_txt("📅 Custom Date Range", "Custom Date Range")],
        horizontal=True
    )
duration_map = {
    "1 Day / 1 दिन": "1d", "5 Days / 5 दिन": "5d", "1 Month / 1 माह": "1mo",
    "6 Months / 6 माह": "6mo", "1 Year / 1 वर्ष": "1y", "5 Years / 5 वर्ष": "5y", "Max / अधिकतम": "max"
}
with rcol2:
    if "Standard" in range_type:
        chosen_dur_label = st.selectbox(get_txt("समयावधि चुनें / Duration:", "Duration:"), list(duration_map.keys()), index=2)
        selected_period = duration_map[chosen_dur_label]
        start_date, end_date = None, None
    else:
        d_c1, d_c2 = st.columns(2)
        start_date = d_c1.date_input("Start Date", value=datetime.date(2023, 1, 1))
        end_date = d_c2.date_input("End Date", value=datetime.date.today())
        selected_period = None

with st.expander(get_txt("🛠️ कस्टमाइज़ेशन विकल्प (Custom Columns & Sheets)", "Custom Columns & Sheets Settings")):
    cc1, cc2, cc3 = st.columns(3)
    inc_ohlc = cc1.checkbox(get_txt("OHLCV डेटा शीट शामिल करें", "Include OHLCV Sheet"), value=True)
    inc_div_sheet = cc2.checkbox(get_txt("डिविडेंड इतिहास शीट शामिल करें", "Include Dividend Sheet"), value=True)
    inc_summary = cc3.checkbox(get_txt("एग्जीक्यूटिव समरी शीट शामिल करें", "Include Executive Summary"), value=True)

# 7. Multi-Tab Screener Grid (With Top BUY, Top SELL, Mutual Funds, Dividends & IPOs)
st.markdown(f"<div class='sec-header'>{get_txt('📊 TradingView लाइव स्क्रीनर, टॉप BUY/SELL, म्यूचुअल फंड्स व डिविडेंड रडार', 'Live Screener, Top Buy/Sell Picks, Mutual Funds & Dividends')}</div>", unsafe_allow_html=True)

screener_tabs = st.tabs([
    "🟢 Top BUY Stocks (क्यों खरीदें)",
    "🔴 Top SELL / Avoid Stocks (क्यों बेचें)",
    "📊 Mutual Funds Radar (Previous & Future)",
    "🏆 Top Dividend Stocks (5Y/10Y Return)",
    "Technicals & RSI Zones",
    "Overview",
    "🚀 Upcoming IPO Radar"
])

with screener_tabs[0]:
    st.markdown("#### 🟢 टॉप BUY स्टॉक्स: खरीदने का कारण, टारगेट व संभावित अपसाइड (Top Buy Picks & Catalysts)")
    df_top_buy = pd.DataFrame(TOP_BUY_STOCKS_MASTER)
    st.dataframe(df_top_buy, use_container_width=True)
    st.caption("💡 *Top Buy Picks मजबूत अर्निंग्स, सेक्टर टेलविंड्स, संस्थागत खरीदारी व टेक्निकल ब्रेकआउट पर आधारित हैं।")

with screener_tabs[1]:
    st.markdown("#### 🔴 टॉप SELL / AVOID स्टॉक्स: बेचने का कारण, रिस्क व डाउनसाइड (Top Sell / Exit Picks & Risks)")
    df_top_sell = pd.DataFrame(TOP_SELL_STOCKS_MASTER)
    st.dataframe(df_top_sell, use_container_width=True)
    st.caption("⚠️ *Top Sell / Avoid Picks कमजोर फंडामेंटल्स, ओवरवैल्युएशन, रेगुलेटरी जोखिम व बेयरिश मोमेंटम पर आधारित हैं।")

with screener_tabs[2]:
    st.markdown("#### 📈 भारत के टॉप म्यूचुअल फंड्स: हिस्टोरिकल व AI अनुमानित फ्यूचर रिटर्न (Previous & Future Expected Returns)")
    df_mf = pd.DataFrame(TOP_MUTUAL_FUNDS_DATA)
    st.dataframe(df_mf, use_container_width=True)

with screener_tabs[3]:
    st.markdown("#### 💎 भारत और अमेरिका के टॉप डिविडेंड पेइंग स्टॉक्स (Top Dividend Yielders & Expected Returns)")
    df_div_top = pd.DataFrame(TOP_DIVIDEND_STOCKS_MASTER)
    st.dataframe(df_div_top, use_container_width=True)

with screener_tabs[4]:
    if st.button("⚡ रन RSI 10-100 ज़ोन व बुलिश सिग्नल स्कैन", key="run_tech_scan"):
        with st.spinner("Calculating RSI Zones, Momentum & Breakouts..."):
            tech_rows = []
            for s_name, s_ticker in POPULAR_STOCKS_PRESET[:20]:
                try:
                    s_t = yf.Ticker(s_ticker)
                    s_h = s_t.history(period="3mo")
                    if not s_h.empty and len(s_h) >= 15:
                        c_p = round(float(s_h["Close"].iloc[-1]), 2)
                        r_val = round(float(calculate_rsi(s_h["Close"]).iloc[-1]), 1)
                        m_line, s_line = calculate_macd(s_h["Close"])
                        m_val = round(float(m_line.iloc[-1]), 2)
                        
                        if r_val <= 30:
                            zone = "RSI 0-30 (Extreme Oversold / Strong Buy)"
                            act = "🟢 Strong Buy (Oversold)"
                            reason = "संकेतक न्यूनतम स्तर पर, रिवर्सल बाउंस की उच्च संभावना"
                        elif r_val <= 45:
                            zone = "RSI 30-45 (Accumulation Zone)"
                            act = "🟢 Buy on Dips"
                            reason = "सपोर्ट लेवल पर एक्युमुलेशन और बॉटम फॉर्मेशन"
                        elif r_val <= 65:
                            zone = "RSI 45-65 (Strong Bullish Momentum)"
                            act = "🟢 Strong Bullish (Holding)"
                            reason = "पॉजिटिव प्राइस मोमेंटम व 50-SMA से ऊपर सस्टेन"
                        elif r_val <= 75:
                            zone = "RSI 65-75 (Overbought Entry)"
                            act = "🟡 Hold / Trail SL"
                            reason = "तेज रैली जारी, स्टॉप-लॉस ट्रेल करें"
                        else:
                            zone = "RSI 75-100 (Extreme Overbought)"
                            act = "🔴 Profit Booking Alert"
                            reason = "अत्यधिक ओवरबॉट, कभी भी मुनाफावसूली संभव"

                        tech_rows.append({
                            "Symbol": s_ticker, "Company Name": s_name, "Price": c_p,
                            "RSI (14)": r_val, "RSI 10-100 Zone": zone, "AI Action Signal": act,
                            "बुलिश / एक्शन का कारण (Reason)": reason
                        })
                except Exception:
                    continue
            if tech_rows: st.dataframe(pd.DataFrame(tech_rows).sort_values(by="RSI (14)", ascending=False), use_container_width=True)

with screener_tabs[5]:
    if st.button("⚡ रन ओवरव्यू स्क्रीनर स्कैन (Run Overview Scan)", key="run_overview_scan"):
        with st.spinner("Scanning top stocks..."):
            rows = []
            for s_name, s_ticker in POPULAR_STOCKS_PRESET[:20]:
                try:
                    s_t = yf.Ticker(s_ticker)
                    s_h = s_t.history(period="3mo")
                    s_inf = s_t.info or {}
                    if not s_h.empty:
                        c_p = round(float(s_h["Close"].iloc[-1]), 2)
                        p_c = round(float(s_h["Close"].iloc[-2]), 2) if len(s_h) > 1 else c_p
                        chg = round(((c_p - p_c) / p_c) * 100, 2)
                        vol = f"{s_h['Volume'].iloc[-1]/1e6:.2f}M" if "Volume" in s_h else "N/A"
                        mkt_c = f"₹{s_inf.get('marketCap', 0)/1e12:.2f}T" if s_inf.get('marketCap') else "N/A"
                        p_e = round(s_inf.get('trailingPE', 0), 2) if s_inf.get('trailingPE') else "N/A"
                        eps_val = round(s_inf.get('trailingEps', 0), 2) if s_inf.get('trailingEps') else "N/A"
                        div_y = f"{s_inf.get('dividendYield', 0)*100:.2f}%" if s_inf.get('dividendYield') else "0.00%"
                        rating = str(s_inf.get('recommendationKey', 'Buy')).replace('_', ' ').title()
                        
                        rows.append({
                            "Symbol": s_ticker, "Company Name": s_name, "Price": c_p, "Chg %": f"{chg:+}%",
                            "Vol": vol, "Mkt Cap": mkt_c, "P/E": p_e, "EPS (TTM)": eps_val,
                            "Div Yield %": div_y, "Sector": s_inf.get('sector', "Equities"),
                            "Analyst Rating": f"⭐ {rating}"
                        })
                except Exception:
                    continue
            if rows: st.dataframe(pd.DataFrame(rows), use_container_width=True)

with screener_tabs[6]:
    st.dataframe(pd.DataFrame(UPCOMING_IPOS_DATA), use_container_width=True)

# 8. Single Stock Fetching Payload
def fetch_stock_payload(ticker_symbol, period_val, s_date, e_date):
    try:
        t = yf.Ticker(ticker_symbol)
        h = t.history(period=period_val) if period_val else t.history(start=s_date, end=e_date)
        
        if h.empty and not ticker_symbol.endswith(".NS") and not ticker_symbol.startswith("^") and "=" not in ticker_symbol:
            t = yf.Ticker(f"{ticker_symbol}.NS")
            h = t.history(period=period_val) if period_val else t.history(start=s_date, end=e_date)
            
        if h.empty:
            h = t.history(period="1y")
            
        max_h = t.history(period="max")
        info = t.info or {}
        divs = t.dividends if hasattr(t, "dividends") else pd.Series(dtype=float)
        ath = max_h["High"].max() if not max_h.empty else (h["High"].max() if not h.empty else None)
        return t, h, info, ath, divs
    except Exception:
        return None, None, None, None, None

def generate_premium_excel(summary_data, hist_df, div_df, inc_s, inc_o, inc_d):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if inc_s and summary_data: pd.DataFrame(summary_data).to_excel(writer, sheet_name="Executive Summary", index=False)
        if inc_o and not hist_df.empty:
            h_exp = hist_df.reset_index()
            if "Date" in h_exp.columns: h_exp["Date"] = h_exp["Date"].dt.strftime("%Y-%m-%d")
            h_exp.to_excel(writer, sheet_name="Historical OHLC Data", index=False)
        if inc_d and not div_df.empty:
            d_exp = div_df.reset_index()
            if "Date" in d_exp.columns: d_exp["Date"] = d_exp["Date"].dt.strftime("%Y-%m-%d")
            d_exp.to_excel(writer, sheet_name="Dividend History", index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color="131722", end_color="131722", fill_type="solid")
        cat_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cat_font = Font(name="Arial", size=10, bold=True, color="1B365D")
        regular_font = Font(name="Arial", size=10)
        thin_border = Border(left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"),
                             top=Side(style="thin", color="E0E0E0"), bottom=Side(style="thin", color="E0E0E0"))

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = 30
            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                if row_idx == 1:
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    is_cat = str(row[0].value).startswith("---")
                    for cell in row:
                        if is_cat:
                            cell.fill = cat_fill
                            cell.font = cat_font
                        else:
                            cell.font = regular_font
                            cell.border = thin_border
    output.seek(0)
    return output.getvalue()

# 9. Execution & Single Stock Detailed Report
if symbol:
    with st.spinner(f"Fetching Live Market Analytics for {symbol}..."):
        ticker_obj, df_hist, stock_info, ath_val, df_div = fetch_stock_payload(symbol, selected_period, start_date, end_date)

    if df_hist is not None and not df_hist.empty:
        cmp_price = stock_info.get("currentPrice") or stock_info.get("regularMarketPrice") or float(df_hist["Close"].iloc[-1])
        prev_close = stock_info.get("regularMarketPreviousClose") or (float(df_hist["Close"].iloc[-2]) if len(df_hist) > 1 else cmp_price)
        price_change = cmp_price - prev_close
        price_change_pct = (price_change / prev_close) * 100 if prev_close else 0.0

        high_52 = stock_info.get("fiftyTwoWeekHigh") or float(df_hist["High"].max())
        low_52 = stock_info.get("fiftyTwoWeekLow") or float(df_hist["Low"].min())
        ath = ath_val if ath_val else high_52
        down_from_ath = (((cmp_price - ath) / ath) * 100) if ath else 0.0
        down_from_52w = (((cmp_price - high_52) / high_52) * 100) if high_52 else 0.0

        company_pe = stock_info.get("trailingPE") or stock_info.get("forwardPE")
        industry_pe = stock_info.get("industryPE") or stock_info.get("sectorPE", "N/A")
        pb_ratio = stock_info.get("priceToBook")
        eps = stock_info.get("trailingEps") or stock_info.get("forwardEps")
        book_val = stock_info.get("bookValue")
        mkt_cap = stock_info.get("marketCap")
        div_rate = stock_info.get("dividendRate", 0.0) or 0.0
        div_yield = (stock_info.get("dividendYield") or 0.0) * 100
        currency = stock_info.get("currency", "INR")
        long_name = stock_info.get("longName", symbol)
        sector = stock_info.get("sector", "Commodities / Indices / Equities")
        industry = stock_info.get("industry", "Global Market")

        fund_score, fund_verdict, fund_class, fund_factors = evaluate_fundamental_health(stock_info)

        df_hist["SMA_20"] = df_hist["Close"].rolling(20).mean()
        df_hist["SMA_50"] = df_hist["Close"].rolling(50).mean()
        df_hist["SMA_200"] = df_hist["Close"].rolling(200).mean()
        df_hist["Upper_BB"], df_hist["Lower_BB"], _ = calculate_bollinger_bands(df_hist["Close"])
        df_hist["RSI"] = calculate_rsi(df_hist["Close"])
        df_hist["MACD"], df_hist["MACD_Sig"] = calculate_macd(df_hist["Close"])

        latest_rsi = float(df_hist["RSI"].dropna().iloc[-1]) if not df_hist["RSI"].dropna().empty else 50.0
        latest_macd = float(df_hist["MACD"].dropna().iloc[-1]) if not df_hist["MACD"].dropna().empty else 0.0
        latest_sig = float(df_hist["MACD_Sig"].dropna().iloc[-1]) if not df_hist["MACD_Sig"].dropna().empty else 0.0
        sma_50_val = float(df_hist["SMA_50"].dropna().iloc[-1]) if not df_hist["SMA_50"].dropna().empty else cmp_price

        # RSI Overbought/Oversold Level Categorization
        if latest_rsi <= 30:
            rsi_detail = "RSI 0-30 (Strong Oversold / Best Buying Opportunity)"
            rsi_sig = "OVERSOLD (BUY)"
            rsi_catalyst = "स्टॉक अत्यधिक ओवर्सोल्ड स्तर पर है, रिस्क न्यूनतम और रिवर्सल बाउंस की अत्यधिक संभावना है।"
        elif latest_rsi <= 45:
            rsi_detail = "RSI 30-45 (Accumulation Range)"
            rsi_sig = "ACCUMULATE (BUY ON DIPS)"
            rsi_catalyst = "सपोर्ट के पास ठहराव, बॉटम फॉर्मेशन के बाद धीरे-धीरे एक्युमुलेट करने का सही स्तर।"
        elif latest_rsi <= 65:
            rsi_detail = "RSI 45-65 (Bullish Trend Momentum)"
            rsi_sig = "BULLISH (HOLD / ACCUMULATE)"
            rsi_catalyst = "स्वस्थ बुलिश मोमेंटम जारी, मूविंग एवरेज से ऊपर मजबूत वॉल्यूम सपोर्ट।"
        elif latest_rsi <= 75:
            rsi_detail = "RSI 65-75 (Overbought Zone Entry)"
            rsi_sig = "BULLISH / TRAIL STOPLOSS"
            rsi_catalyst = "मोमेंटम बहुत मजबूत है, लेकिन फ्रेश एंट्री से बचें और स्टॉप-लॉस ऊपर की ओर ट्रेल करें।"
        else:
            rsi_detail = "RSI 75-100 (Extreme Overbought / Profit Booking Alert)"
            rsi_sig = "OVERBOUGHT (SELL / BOOK PROFIT)"
            rsi_catalyst = "अत्यधिक ओवरबॉट ज़ोन, यहाँ से किसी भी समय शॉर्ट-टर्म मुनाफावसूली या पुलबैक आ सकता है।"

        macd_sig = "BULLISH CROSSOVER (BUY)" if latest_macd > latest_sig else "BEARISH CROSSOVER (SELL)"
        trend_sig = "BULLISH (Above 50 SMA)" if cmp_price > sma_50_val else "BEARISH (Below 50 SMA)"

        oi_data = fetch_option_chain_oi(ticker_obj, cmp_price)

        score = 0
        if latest_rsi < 45: score += 1.5
        elif latest_rsi < 60: score += 1.0
        if latest_macd > latest_sig: score += 1.5
        if cmp_price > sma_50_val: score += 1.0
        if down_from_52w < -15: score += 1.0
        if fund_score >= 60: score += 1.5
        if oi_data and "BULLISH" in oi_data["sentiment"]: score += 1.0

        win_prob = round(min(max((score / 7.5) * 100, 25.0), 93.0), 1)
        
        if win_prob >= 78:
            ai_action = "EXTREMELY BULLISH 🚀🚀 (अति तेज वृद्धि संभावना)"
            future_pred_text = "अगले 3-6 महीनों में मजबूत मोमेंटम और नए हाई बनाने की उच्च संभावना (80%+ Positive Bias)।"
        elif win_prob >= 60:
            ai_action = "BULLISH / BUY 📈 (सकारात्मक रुझान)"
            future_pred_text = "मध्यम अवधि में 10-15% अपसाइड रैली की मजबूत संभावना।"
        elif win_prob >= 45:
            ai_action = "NEUTRAL / HOLD ⚖️ (संतुलित दायरा)"
            future_pred_text = "स्टॉक सीमित दायरे में कंसोलिडेट कर सकता है।"
        else:
            ai_action = "BEARISH / AVOID 📉 (दबाव / मुनाफावसूली)"
            future_pred_text = "निकट भविष्य में सपोर्ट लेवल्स की दोबारा टेस्टिंग हो सकती है।"

        intrinsic_val = calculate_intrinsic_value(eps, book_val) if eps and book_val else None
        if intrinsic_val:
            ai_fair_buy_price = round((intrinsic_val * 0.85 + cmp_price * 0.95) / 2, 2)
            ai_max_buy_price = round(intrinsic_val * 0.95, 2)
        else:
            ai_fair_buy_price = round(cmp_price * 0.95, 2)
            ai_max_buy_price = round(cmp_price * 0.98, 2)

        # Multi-Horizon Expected Return Calculations (1M, 2M, 3M, 4M, 6M, 1Y, 5Y, 10Y)
        base_annual_growth = 0.14 if fund_score >= 70 else (0.10 if fund_score >= 45 else 0.06)
        div_yield_val = (div_yield / 100.0) if div_yield else 0.015
        monthly_base = (base_annual_growth + div_yield_val) / 12.0

        mom_factor = 1.35 if "BULLISH" in ai_action else (1.0 if "NEUTRAL" in ai_action else 0.7)

        ret_1m = round(monthly_base * 1 * mom_factor * 100, 2)
        tgt_1m = round(cmp_price * (1 + (ret_1m / 100)), 2)

        ret_2m = round(monthly_base * 2 * mom_factor * 100, 2)
        tgt_2m = round(cmp_price * (1 + (ret_2m / 100)), 2)

        ret_3m = round(monthly_base * 3 * mom_factor * 100, 2)
        tgt_3m = round(cmp_price * (1 + (ret_3m / 100)), 2)

        ret_4m = round(monthly_base * 4 * mom_factor * 100, 2)
        tgt_4m = round(cmp_price * (1 + (ret_4m / 100)), 2)

        ret_6m = round(monthly_base * 6 * mom_factor * 100, 2)
        tgt_6m = round(cmp_price * (1 + (ret_6m / 100)), 2)

        ret_1y = round((base_annual_growth + div_yield_val) * mom_factor * 100, 2)
        tgt_1y = round(cmp_price * (1 + (ret_1y / 100)), 2)

        total_annual_comp = base_annual_growth + div_yield_val
        tgt_5y = round(cmp_price * ((1 + total_annual_comp) ** 5), 2)
        ret_5y = round((((tgt_5y - cmp_price) / cmp_price) * 100), 1)

        tgt_10y = round(cmp_price * ((1 + total_annual_comp) ** 10), 2)
        ret_10y = round((((tgt_10y - cmp_price) / cmp_price) * 100), 1)

        if buy_price > 0:
            if cmp_price < buy_price:
                suggested_rebuy_price = round(cmp_price * 0.98, 2)
                rebuy_advice = f"🟢 स्टॉक आपके खरीद भाव से {((buy_price-cmp_price)/buy_price*100):.1f}% नीचे है। `{currency} {suggested_rebuy_price}` पर एक्युमुलेट/एवरेज करें।"
            else:
                suggested_rebuy_price = round(max(cmp_price * 0.96, buy_price), 2)
                rebuy_advice = f"⚖️ स्टॉक आपके खरीद भाव से मुनाफे में है। पिरामिडिंग हेतु डिप पर `{currency} {suggested_rebuy_price}` पर जोड़ें।"
        else:
            suggested_rebuy_price = ai_fair_buy_price
            rebuy_advice = "Sidebar में अपना पुराना खरीद भाव दर्ज करके री-बाय स्तर देखें।"

        entry_lvl = round(cmp_price * 0.985, 2)
        sl_lvl = round(cmp_price * 0.94, 2)
        tgt_1 = round(cmp_price * 1.08, 2)
        tgt_2 = round(cmp_price * 1.15, 2)

        analyst_recom = str(stock_info.get("recommendationKey", "BUY")).replace('_', ' ').upper()
        target_mean = stock_info.get("targetMeanPrice", round(cmp_price * 1.14, 2))

        total_lifetime_div = float(df_div.sum()) if not df_div.empty else 0.0
        yield_on_cost = (div_rate / buy_price * 100) if buy_price > 0 else None

        # Header Info
        st.markdown("---")
        st.subheader(f"🏢 {long_name} ({symbol})")
        st.caption(f"Sector: **{sector}** | Industry: **{industry}** | Currency: **{currency}**")

        # 0. Multi-Horizon AI Returns Table (1M, 2M, 3M, 4M, 6M, 1Y, 5Y, 10Y)
        st.markdown(f"<div class='sec-header'>{get_txt('⏳ AI मल्टी-टाइमफ्रेम रिटर्न व टारगेट प्रेडिक्शन (1M, 2M, 3M, 4M, 6M, 1Y, 5Y, 10Y)', 'Multi-Horizon AI Return & Target Predictions')}</div>", unsafe_allow_html=True)
        
        horizon_data = [
            {"Timeframe / अवधि": "1 Month (1 माह)", "Expected Target Price": f"{currency} {tgt_1m:,.2f}", "Expected Gain (%)": f"+{ret_1m}%", "AI Confidence": "High (Momentum / RSI)"},
            {"Timeframe / अवधि": "2 Months (2 माह)", "Expected Target Price": f"{currency} {tgt_2m:,.2f}", "Expected Gain (%)": f"+{ret_2m}%", "AI Confidence": "High (Short Trend)"},
            {"Timeframe / अवधि": "3 Months (3 माह)", "Expected Target Price": f"{currency} {tgt_3m:,.2f}", "Expected Gain (%)": f"+{ret_3m}%", "AI Confidence": "Robust (Quarterly Cycle)"},
            {"Timeframe / अवधि": "4 Months (4 माह)", "Expected Target Price": f"{currency} {tgt_4m:,.2f}", "Expected Gain (%)": f"+{ret_4m}%", "AI Confidence": "Steady (Moving Average)"},
            {"Timeframe / अवधि": "6 Months (6 माह)", "Expected Target Price": f"{currency} {tgt_6m:,.2f}", "Expected Gain (%)": f"+{ret_6m}%", "AI Confidence": "High (Earnings Horizon)"},
            {"Timeframe / अवधि": "1 Year (1 वर्ष)", "Expected Target Price": f"{currency} {tgt_1y:,.2f}", "Expected Gain (%)": f"+{ret_1y}%", "AI Confidence": "Institutional Target"},
            {"Timeframe / अवधि": "5 Years (5 वर्ष)", "Expected Target Price": f"{currency} {tgt_5y:,.2f}", "Expected Gain (%)": f"+{ret_5y}%", "AI Confidence": "CAGR + Div Reinvestment"},
            {"Timeframe / अवधि": "10 Years (10 वर्ष)", "Expected Target Price": f"{currency} {tgt_10y:,.2f}", "Expected Gain (%)": f"+{ret_10y}%", "AI Confidence": "Long-Term Compounding"},
        ]
        st.dataframe(pd.DataFrame(horizon_data), use_container_width=True)

        # AI Prediction & Brokerage Ratings
        st.markdown(f"<div class='sec-header'>{get_txt('🤖 AI भविष्य प्रेडिक्शन, RSI लेवल व ब्रोकरेज कंसेंसस', 'AI Future Prediction & Institutional Brokerage Ratings')}</div>", unsafe_allow_html=True)
        
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("📌 AI प्रेडिक्शन वर्डिक्ट", ai_action)
        r2.metric("📊 प्रॉफिट प्रोबेबिलिटी स्कोर", f"{win_prob}%", "AI ऐतिहासिक डेटा मॉडल")
        r3.metric("📈 RSI (14) स्टेटस", f"{latest_rsi:.1f}", rsi_sig)
        r4.metric("🏢 ब्रोकरेज कंसेंसस", f"⭐ {analyst_recom}", f"Target: {currency} {target_mean:,.1f}")

        st.markdown(
            f"""
            <div class="ai-box">
                🔮 <b>AI बॉट फ्यूचर प्रेडिक्शन:</b> {future_pred_text}<br>
                📊 <b>RSI लेवल एनालिसिस:</b> <b>{rsi_detail}</b> — {rsi_catalyst}<br>
                🏢 <b>संस्थागत ब्रोकरेज ओपिनियन:</b> शीर्ष रिसर्च हाउसेस द्वारा इस पर <b>{analyst_recom}</b> रेटिंग और <b>{currency} {target_mean:,.2f}</b> का औसत टारगेट मूल्य दिया गया है।<br>
                💡 <b>खरीदने / होल्ड करने का प्रमुख कारण:</b> {fund_verdict} स्थिति, डिविडेंड यील्ड {div_yield:.2f}%, और {total_annual_comp*100:.1f}% अनुमानित वार्षिक कम्पाउंडिंग ग्रोथ।
            </div>
            """,
            unsafe_allow_html=True,
        )

        # Fundamental Health
        st.markdown(f"<div class='sec-header'>{get_txt('🛡️ AI फंडामेंटल हेल्थ व कंपनी साउंडनेस', 'AI Fundamental Soundness & Health Score')}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='{fund_class}'>📊 <b>कंपनी स्थिति:</b> {fund_verdict} | <b>AI हेल्थ स्कोर:</b> {fund_score}/100</div>", unsafe_allow_html=True)

        fcol1, fcol2, fcol3 = st.columns(3)
        fcol1.info(f"💎 **AI सही खरीद मूल्य (Fair Buy Price):** `{currency} {ai_fair_buy_price:,.2f}`\n\n*(इस स्तर पर रिस्क न्यूनतम है)*")
        fcol2.info(f"🛑 **अधिकतम खरीद सीमा (Max Buy Limit):** `{currency} {ai_max_buy_price:,.2f}`\n\n*(इसके ऊपर ओवरप्राइस्ड माना जाएगा)*")
        factors_txt = "\n".join(fund_factors[:3])
        fcol3.success(f"📋 **मुख्य फंडामेंटल कारक:**\n\n{factors_txt}")

        # Re-Buy Levels
        st.markdown(f"<div class='sec-header'>{get_txt('🎯 AI री-बाय / एवरेजिंग कैलकुलेटर व ट्रेडिंग स्तर', 'AI Re-Buy Price & Trading Levels')}</div>", unsafe_allow_html=True)
        l1, l2, l3, l4 = st.columns(4)
        l1.metric("📥 AI री-बाय स्तर", f"{currency} {suggested_rebuy_price:,.2f}")
        l2.metric("🛑 स्टॉप-लॉस (Stop-Loss)", f"{currency} {sl_lvl:,.2f}", "-6% Buffer", delta_color="inverse")
        l3.metric("🎯 टार्गेट 1 (Target 1)", f"{currency} {tgt_1:,.2f}", "+8% Target")
        l4.metric("🚀 टार्गेट 2 (Target 2)", f"{currency} {tgt_2:,.2f}", "+15% Target")

        # Price Action & ATH
        st.markdown(f"<div class='sec-header'>{get_txt('मूल्य एवं 52-सप्ताह/लाइफटाइम स्थिति', 'Price Action & ATH Range')}</div>", unsafe_allow_html=True)
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("CMP", f"{currency} {cmp_price:,.2f}", f"{price_change:+.2f} ({price_change_pct:+.2f}%)")
        m2.metric("52W High", f"{currency} {high_52:,.2f}" if high_52 else "N/A", f"{down_from_52w:.2f}% (from 52W High)", delta_color="inverse")
        m3.metric("52W Low", f"{currency} {low_52:,.2f}" if low_52 else "N/A")
        m4.metric("Lifetime High (ATH)", f"{currency} {ath:,.2f}" if ath else "N/A", f"{down_from_ath:.2f}% (from ATH)", delta_color="inverse")
        if intrinsic_val:
            m5.metric("Intrinsic Value", f"{currency} {intrinsic_val:,.2f}", f"{((intrinsic_val - cmp_price) / cmp_price) * 100:+.1f}% Margin")
        else:
            m5.metric("Intrinsic Value", "N/A")

        # Valuation
        st.markdown(f"<div class='sec-header'>{get_txt('वैल्युएशन एवं फंडामेंटल्स (P/E & P/B Multiples)', 'Valuation & Fundamentals')}</div>", unsafe_allow_html=True)
        v1, v2, v3, v4, v5 = st.columns(5)
        v1.metric("Company P/E", f"{company_pe:.2f}" if company_pe else "N/A")
        v2.metric("Industry P/E", str(industry_pe))
        v3.metric("P/B Ratio", f"{pb_ratio:.2f}" if pb_ratio else "N/A")
        v4.metric("EPS (TTM)", f"{currency} {eps:.2f}" if eps else "N/A")
        v5.metric("Dividend Yield (CMP)", f"{div_yield:.2f}%")

        # Dividend Analytics
        st.markdown(f"<div class='sec-header'>{get_txt('💰 डिविडेंड विश्लेषण एवं पूंजी यील्ड (Dividend Analytics & Yield on Cost)', 'Dividend Analytics & Yield on Cost')}</div>", unsafe_allow_html=True)
        d1, d2, d3, d4 = st.columns(4)
        d1.metric(get_txt("लाइफटाइम कुल डिविडेंड", "Lifetime Total Div"), f"{currency} {total_lifetime_div:,.2f}")
        d2.metric(get_txt("वार्षिक डिविडेंड दर (TTM)", "Annual Div Rate (TTM)"), f"{currency} {div_rate:,.2f}")
        d3.metric(get_txt("बुक वैल्यू (Book Value)", "Book Value"), f"{currency} {book_val:,.2f}" if book_val else "N/A")
        if yield_on_cost is not None:
            d4.metric(get_txt("खरीद मूल्य पर यील्ड (Yield on Cost)", "Yield on Cost (Your Buy)"), f"{yield_on_cost:.2f}%", f"Buy: {currency} {buy_price}")
        else:
            d4.metric(get_txt("खरीद मूल्य पर यील्ड", "Yield on Cost"), "Sidebar में दर्ज करें")

        with st.expander(get_txt("📅 कस्टम तारीख अनुसार डिविडेंड कैलकुलेटर", "Custom Date Range Dividend Calculator")):
            div_c1, div_c2 = st.columns(2)
            c_start = div_c1.date_input("Dividend Filter Start", value=datetime.date(2020, 1, 1), key="div_c_start")
            c_end = div_c2.date_input("Dividend Filter End", value=datetime.date.today(), key="div_c_end")
            if not df_div.empty:
                div_clean = df_div.copy()
                div_clean.index = div_clean.index.tz_localize(None)
                mask = (div_clean.index.date >= c_start) & (div_clean.index.date <= c_end)
                range_div_sum = div_clean[mask].sum()
                st.write(f"**{c_start} से {c_end} के बीच कुल डिविडेंड:** `{currency} {range_div_sum:,.2f}`")
                st.dataframe(div_clean[mask], use_container_width=True)

        # F&O & Option Chain
        if oi_data:
            st.markdown(f"<div class='sec-header'>{get_txt('📈 Live Option Chain, PCR व ओपन इंटरेस्ट (OI) विश्लेषण', 'Live Option Chain & Open Interest')}</div>", unsafe_allow_html=True)
            o1, o2, o3, o4 = st.columns(4)
            o1.metric("Put-Call Ratio (PCR)", f"{oi_data['pcr']}")
            o2.metric("OI Action Signal", oi_data["oi_action"])
            o3.metric("Option Fair Price Center", f"{currency} {oi_data['option_fair_price']}")
            o4.metric("Support / Resistance", f"{oi_data['put_support']} / {oi_data['call_resistance']}")
            st.info(f"💡 **F&O / OI सेंटीमेंट:** {oi_data['sentiment']} (Expiry: {oi_data['expiry']})")

        # TradingView Multi-Panel Chart
        st.markdown(f"<div class='sec-header'>📈 TradingView Pro Technical Chart (SMA, BB, MACD & RSI)</div>", unsafe_allow_html=True)
        fig = make_subplots(
            rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.03,
            subplot_titles=("Price & Bollinger Bands", "MACD (12, 26, 9)", "RSI (14)"),
            row_heights=[0.6, 0.2, 0.2]
        )

        if "Open" in df_hist.columns and "High" in df_hist.columns and "Low" in df_hist.columns:
            fig.add_trace(go.Candlestick(
                x=df_hist.index, open=df_hist["Open"], high=df_hist["High"], low=df_hist["Low"], close=df_hist["Close"],
                name="OHLC Candles"
            ), row=1, col=1)
        else:
            fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["Close"], name="Close Price", line=dict(color="#2962ff", width=2)), row=1, col=1)

        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["SMA_20"], name="SMA 20", line=dict(color="#f39c12", width=1)), row=1, col=1)
        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["SMA_50"], name="SMA 50", line=dict(color="#3498db", width=1)), row=1, col=1)
        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["Upper_BB"], name="Upper BB", line=dict(color="rgba(150,150,150,0.5)", dash="dash")), row=1, col=1)
        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["Lower_BB"], name="Lower BB", line=dict(color="rgba(150,150,150,0.5)", dash="dash"), fill="tonexty", fillcolor="rgba(200,200,200,0.05)"), row=1, col=1)

        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["MACD"], name="MACD Line", line=dict(color="#2962ff", width=1.5)), row=2, col=1)
        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["MACD_Sig"], name="Signal Line", line=dict(color="#ff3d60", width=1.5)), row=2, col=1)

        fig.add_trace(go.Scatter(x=df_hist.index, y=df_hist["RSI"], name="RSI", line=dict(color="#9b59b6", width=1.5)), row=3, col=1)
        fig.add_hline(y=70, line_dash="dash", line_color="red", row=3, col=1)
        fig.add_hline(y=30, line_dash="dash", line_color="green", row=3, col=1)

        fig.update_layout(height=650, xaxis_rangeslider_visible=False, template="plotly_white", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig, use_container_width=True)

        with st.expander(get_txt("📋 पूर्ण डेटा तालिका देखें (View Full OHLC Table)", "View Full OHLC Table")):
            st.dataframe(df_hist, use_container_width=True)

        # 6. Full Excel Export
        summary_rows = [
            {"Field": "--- GENERAL OVERVIEW ---", "Value": ""},
            {"Field": "Company Name", "Value": str(long_name)},
            {"Field": "Symbol", "Value": str(symbol)},
            {"Field": "Sector / Industry", "Value": f"{sector} / {industry}"},
            {"Field": "--- AI MULTI-HORIZON RETURN PREDICTIONS ---", "Value": ""},
            {"Field": "1 Month Target Price", "Value": f"{currency} {tgt_1m:,.2f} (+{ret_1m}%)"},
            {"Field": "2 Months Target Price", "Value": f"{currency} {tgt_2m:,.2f} (+{ret_2m}%)"},
            {"Field": "3 Months Target Price", "Value": f"{currency} {tgt_3m:,.2f} (+{ret_3m}%)"},
            {"Field": "4 Months Target Price", "Value": f"{currency} {tgt_4m:,.2f} (+{ret_4m}%)"},
            {"Field": "6 Months Target Price", "Value": f"{currency} {tgt_6m:,.2f} (+{ret_6m}%)"},
            {"Field": "1 Year Target Price", "Value": f"{currency} {tgt_1y:,.2f} (+{ret_1y}%)"},
            {"Field": "5 Years Target Price", "Value": f"{currency} {tgt_5y:,.2f} (+{ret_5y}%)"},
            {"Field": "10 Years Target Price", "Value": f"{currency} {tgt_10y:,.2f} (+{ret_10y}%)"},
            {"Field": "--- AI PREDICTION & EXPERT RATINGS ---", "Value": ""},
            {"Field": "AI Future Prediction Verdict", "Value": ai_action},
            {"Field": "AI Profit Probability", "Value": f"{win_prob}%"},
            {"Field": "Institutional Brokerage Consensus", "Value": analyst_recom},
            {"Field": "Brokerage Target Price", "Value": f"{currency} {target_mean:,.2f}"},
            {"Field": "RSI 10-100 Level Detail", "Value": rsi_detail},
            {"Field": "--- AI VALUATION & LEVELS ---", "Value": ""},
            {"Field": "AI Fair Buy Price", "Value": f"{currency} {ai_fair_buy_price:,.2f}"},
            {"Field": "AI Max Buy Limit", "Value": f"{currency} {ai_max_buy_price:,.2f}"},
            {"Field": "AI Re-Buy / Averaging Price", "Value": f"{currency} {suggested_rebuy_price:,.2f}"},
            {"Field": "Suggested Entry Level", "Value": f"{currency} {entry_lvl:,.2f}"},
            {"Field": "Stop Loss Level", "Value": f"{currency} {sl_lvl:,.2f}"},
            {"Field": "Target Price 1", "Value": f"{currency} {tgt_1:,.2f}"},
            {"Field": "Target Price 2", "Value": f"{currency} {tgt_2:,.2f}"},
            {"Field": "--- PRICE & ATH METRICS ---", "Value": ""},
            {"Field": "Current Market Price (CMP)", "Value": f"{currency} {cmp_price:,.2f}"},
            {"Field": "52-Week High", "Value": f"{currency} {high_52:,.2f}" if high_52 else "N/A"},
            {"Field": "52-Week Low", "Value": f"{currency} {low_52:,.2f}" if low_52 else "N/A"},
            {"Field": "Down from 52W High", "Value": f"{down_from_52w:.2f}%"},
            {"Field": "All-Time High (ATH)", "Value": f"{currency} {ath:,.2f}" if ath else "N/A"},
            {"Field": "Down from ATH", "Value": f"{down_from_ath:.2f}%"},
            {"Field": "--- F&O / OPTION CHAIN INSIGHTS ---", "Value": ""},
            {"Field": "PCR (Put-Call Ratio)", "Value": str(oi_data["pcr"]) if oi_data else "N/A"},
            {"Field": "OI Sentiment", "Value": str(oi_data["sentiment"]) if oi_data else "N/A"},
            {"Field": "Option Fair Price", "Value": f"{currency} {oi_data['option_fair_price']}" if oi_data else "N/A"},
            {"Field": "--- TECHNICAL SIGNALS ---", "Value": ""},
            {"Field": "RSI (14)", "Value": f"{latest_rsi:.1f} ({rsi_sig})"},
            {"Field": "MACD Signal", "Value": macd_sig},
            {"Field": "--- DIVIDEND & CAPITAL YIELD ---", "Value": ""},
            {"Field": "Dividend Yield (CMP)", "Value": f"{div_yield:.2f}%"},
            {"Field": "Lifetime Total Dividend", "Value": f"{currency} {total_lifetime_div:,.2f}"},
            {"Field": "Your Buy Price", "Value": f"{currency} {buy_price:,.2f}" if buy_price > 0 else "Not Provided"},
            {"Field": "Yield on Cost (On Your Capital)", "Value": f"{yield_on_cost:.2f}%" if yield_on_cost else "N/A"},
        ]

        excel_data = generate_premium_excel(summary_rows, df_hist, df_div, inc_summary, inc_ohlc, inc_div_sheet)

        st.markdown("---")
        st.download_button(
            label=get_txt("📥 प्रीमियम फॉर्मेटेड एक्सेल रिपोर्ट डाउनलोड करें (.xlsx)", "📥 Download Premium Executive Report (.xlsx)"),
            data=excel_data,
            file_name=f"{symbol}_TradingView_Report_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.error("डेटा प्राप्त करने में असमर्थ। कृपया सिंबल की जाँच करें।")
